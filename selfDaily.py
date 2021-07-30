#PythonとExcelを使用して日報を書けるようにしたい
import openpyxl as px
import PySimpleGUI as sg
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
#Excelの読み込み
filename = 'daily.xlsx'
wb = px.load_workbook(filename)
ws = wb['Sheet3']

#GUIのカレンダーから時刻データをなくす関数
def read_date():
    _cal = values['-cal-']
    cal = datetime.strptime(_cal, '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d')
    return cal

#GUIを作成
sg.theme('TanBlue')
layout = [
    [sg.Text('日付を選択してください')],
    [sg.CalendarButton('カレンダー', target='-cal-')],
    [sg.Input(size=(45, 2), key='-cal-')],
    [sg.Text('１日の業務を入力してください')],
    [sg.Multiline(key='-in-')],
    [sg.Submit('登録', size=(10,1), key='-go-'),
     sg.Button('中止', size=(10, 1), key='-q-')]
    ]

layout2 = [
    #[sg.Text('修正する場合は以下を入力してください', key='-re-')],
    [sg.CalendarButton('修正カレンダー', target='-rcal-')],
    [sg.Input(size=(45,2), key='-rcal-')],
    [sg.Text('修正する業務を入力してください')],
    [sg.Multiline(key='-rin-')],
    [sg.Submit('再登録', size=(10,1), key='-rgo-'),
     sg.Button('中止', size=(10, 1), key='-rq-')]
    ]

layout3 = [
    [sg.Frame('入力欄',layout)],
    [sg.Frame('修正する場合は以下を入力してください',
              layout2,title_location = sg.TITLE_LOCATION_TOP_LEFT)]
    ]

#ウインドウの読み込み
window = sg.Window('業務日報', layout3)

while True:
    event, values = window.read()
    if event in (None, '-q-', '-rq-'):
        break

    if event == '-go-':
        if  values['-cal-'] == '' or len(values['-in-']) == 1:
            print('!',len(values['-in-']))
            sg.popup('情報を正しく入力してください', title='Attention')
        else:
            #print('ok', ws.cell(row=ws.max_row+1, column=1).value) 
            ws.cell(row=ws.max_row+1, column=1).value = read_date()
            ws.cell(row=ws.max_row, column=2).value = values['-in-']
            wb.save(filename)
            sg.popup('登録しました', title='お疲れ様でした')


    elif event == '-rgo-':
        if  values['-rcal-'] == '' or len(values['-rin-']) == 1:
            sg.popup('情報を正しく入力してください', title='Attention!')
        else:
            _rcal = values['-rcal-']
            rcal = datetime.strptime(_rcal, '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d')
            for i in range(3, ws.max_row+1):
                if ws.cell(row=i, column=1).value == rcal:
                    ws.cell(row=i, column=2).value = values['-rin-']
                    wb.save(filename)
                    sg.popup('再登録しました', title='お知らせ')
                    break
            else:
                sg.popup('日付はみつかりません', title='Attentione!')


